import openpyxl as xl
book = xl.load_workbook('PriceSheet.xlsx')
sheet = book['Sheet1']
cell = sheet.cell(1, 1)
print(sheet.max_row)
